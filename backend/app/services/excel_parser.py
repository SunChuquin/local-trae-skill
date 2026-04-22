"""
Excel 解析与清洗引擎
支持 xlsx/xls 全格式，多 Sheet 批量读取，自动处理合并单元格、公式、空值等
"""
import os
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
from datetime import datetime
import uuid

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from loguru import logger
from app.models.excel_document import SheetInfo, ParsePreview


class ExcelParser:
    """
    Excel 文档解析器
    支持 xlsx/xls 格式，提供数据清洗、Sheet 信息提取、表格预览等功能
    """

    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xls']
        self.max_preview_rows = 100

    def validate_file(self, file_path: str) -> Tuple[bool, str]:
        """
        验证文件是否有效
        返回: (是否有效, 错误信息)
        """
        path = Path(file_path)
        if not path.exists():
            return False, "文件不存在"
        if path.suffix.lower() not in self.supported_extensions:
            return False, f"不支持的文件格式: {path.suffix}"
        try:
            wb = load_workbook(file_path, data_only=True)
            wb.close()
            return True, ""
        except InvalidFileException as e:
            return False, f"文件格式错误或已损坏: {str(e)}"
        except Exception as e:
            return False, f"文件读取失败: {str(e)}"

    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """
        提取文件元数据
        """
        path = Path(file_path)
        return {
            "name": path.name,
            "size": path.stat().st_size,
            "size_formatted": self._format_size(path.stat().st_size),
            "extension": path.suffix.lower()
        }

    def get_sheets_info(self, file_path: str) -> List[SheetInfo]:
        """
        获取所有 Sheet 的基本信息
        """
        try:
            wb = load_workbook(file_path, data_only=True)
            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheets.append(SheetInfo(
                    name=sheet_name,
                    row_count=ws.max_row,
                    col_count=ws.max_column,
                    merged_cells=len(ws.merged_cells.ranges)
                ))
            wb.close()
            return sheets
        except Exception as e:
            logger.error(f"获取 Sheet 信息失败: {str(e)}")
            return []

    def parse_sheet_data(
        self,
        file_path: str,
        sheet_name: str,
        include_merged_values: bool = True
    ) -> Tuple[List[str], List[List[str]]]:
        """
        解析指定 Sheet 的数据
        返回: (表头列表, 数据行列表)
        """
        try:
            wb = load_workbook(file_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return [], []
            ws = wb[sheet_name]

            headers = []
            rows = []

            merged_cells_values = {}
            if include_merged_values:
                for merged_range in ws.merged_cells.ranges:
                    min_col = merged_range.min_col
                    min_row = merged_range.min_row
                    max_col = merged_range.max_col
                    max_row = merged_range.max_row
                    cell_value = ws.cell(min_row, min_col).value
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            merged_cells_values[(row, col)] = cell_value

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    headers = [self._clean_value(cell) for cell in row]
                    headers = [h if h else f"列{col_idx}" for col_idx, h in enumerate(headers, 1)]
                else:
                    row_values = []
                    for col_idx, cell in enumerate(row, start=1):
                        key = (row_idx, col_idx)
                        if key in merged_cells_values:
                            row_values.append(self._clean_value(merged_cells_values[key]))
                        else:
                            row_values.append(self._clean_value(cell))
                    if self._is_valid_row(row_values, len(headers)):
                        rows.append(row_values)

            wb.close()
            return headers, rows

        except Exception as e:
            logger.error(f"解析 Sheet {sheet_name} 失败: {str(e)}")
            return [], []

    def parse_all_sheets(self, file_path: str) -> Dict[str, Tuple[List[str], List[List[str]]]]:
        """
        解析所有 Sheet
        返回: {sheet_name: (表头, 数据行)}
        """
        result = {}
        sheets_info = self.get_sheets_info(file_path)
        for sheet in sheets_info:
            headers, rows = self.parse_sheet_data(file_path, sheet.name)
            result[sheet.name] = (headers, rows)
        return result

    def get_preview(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        max_rows: int = 50
    ) -> List[ParsePreview]:
        """
        获取表格预览数据
        """
        previews = []
        sheets_to_preview = [sheet_name] if sheet_name else [s.name for s in self.get_sheets_info(file_path)]

        for sname in sheets_to_preview[:5]:
            headers, rows = self.parse_sheet_data(file_path, sname)
            preview_rows = rows[:max_rows]
            previews.append(ParsePreview(
                sheet_name=sname,
                headers=headers,
                rows=preview_rows,
                total_rows=len(rows),
                total_cols=len(headers)
            ))
        return previews

    def _clean_value(self, value: Any) -> str:
        """
        清洗单元格值
        """
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            if pd.isna(value):
                return ""
            return str(int(value)) if float(value).is_integer() else str(round(value, 4))
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return str(value).strip()

    def _is_valid_row(self, row: List[str], expected_cols: int) -> bool:
        """
        判断是否为有效数据行（去除全空行、无效行）
        """
        if len(row) == 0:
            return False
        non_empty = sum(1 for v in row if v and v.strip())
        return non_empty >= expected_cols * 0.3

    def _format_size(self, size: int) -> str:
        """
        格式化文件大小
        """
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size < 1024:
                return f"{size:.2f} {unit}"
            size /= 1024
        return f"{size:.2f} TB"


excel_parser = ExcelParser()
